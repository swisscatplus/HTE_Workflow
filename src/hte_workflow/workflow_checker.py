import argparse
from typing import Dict, List, Tuple, Any, Optional, Iterable
from pathlib import Path
import re
import sys
import json
import math
import matplotlib.pyplot as plt

import openpyxl
import pandas as pd
import numpy as np
# from matplotlib.sphinxext.plot_directive import out_of_date

try:
    import seaborn as sns
    _HAS_SNS = True
except Exception:
    _HAS_SNS = False

from hte_workflow import layout_parser as lp
from hte_workflow.paths import DATA_DIR, OUT_DIR

"""
Labelling
"""
# ==== LABEL NORMALIZATION ====
_UNIT_PARENS = re.compile(r"\s*\((?:u?l|µl|ml|mg|g|mm|mM|M)\)\s*$", re.IGNORECASE)
_IN_CLAUSE = re.compile(r"\s+in\s+.+$", re.IGNORECASE)
_WS = re.compile(r"\s+")

def _norm_label(s: str) -> str:
    if s is None:
        return ""
    s0 = str(s)
    s1 = _IN_CLAUSE.sub("", s0)             # drop " in DMF" tails
    s2 = _UNIT_PARENS.sub("", s1)           # drop trailing "(uL)" etc
    s3 = _WS.sub(" ", s2).strip().lower()   # collapse ws, lowercase
    return s3

def _load_label_map(csv_path: Optional[Path]) -> Dict[str, str]:
    """
    Optional CSV with columns: workflow_label,synthesis_name
    Keys are normalized workflow labels; values are used as exact synthesis column names.
    """
    if not csv_path:
        return {}
    if not csv_path.exists():
        print(f"[label-map] Not found: {csv_path} (ignored)", file=sys.stderr)
        return {}
    df = pd.read_csv(csv_path)
    out = {}
    for _, row in df.iterrows():
        wf = str(row.get("workflow_label", "")).strip()
        sy = str(row.get("synthesis_name", "")).strip()
        if wf and sy:
            out[_norm_label(wf)] = sy
    print(f"[label-map] loaded {len(out)} overrides.", file=sys.stderr)
    return out

"""
---------- Parsing synthesis JSON to per-experiment DataFrame ----------
"""
_SOLVENT_GROUPS = {"solvent", "colvents"}            # 'colvents' = co-solvents group
_LIQUID_STATES  = {"liquid", "Liquid", "LIQUID"}

def _load_synthesis(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def _experiment_order(meta: Dict[str, Any], exps: Dict[str, Any]) -> List[str]:
    m = meta.get("experiment_to_well") or []
    if m:
        return [str(x["experiment"]) for x in m]
    # fallback to numeric sort
    return sorted(exps.keys(), key=lambda z: int(str(z)))

def _per_experiment_from_synthesis(
    synth: Dict[str, Any],
) -> Tuple[pd.DataFrame, Dict[str, Any], List[str]]:
    """
    Returns:
      df: index 'Experiment 1..N', columns '<ChemicalName> (uL|mg)'
      info: dict with 'V_target' and other meta
      warnings: list of human-readable warnings
    """
    meta = synth.get("meta", {})
    exps = synth.get("experiments", {})
    order = _experiment_order(meta, exps)

    # Target well volume (uL)
    V_target = None
    lim = meta.get("limiting", {})
    if "well_volume_uL" in lim:
        try:
            V_target = float(lim["well_volume_uL"])
        except Exception:
            V_target = None

    rows: List[Tuple[str, Dict[str, float]]] = []
    warns: List[str] = []

    for i, eid in enumerate(order, start=1):
        dispenses = exps.get(str(eid), {}).get("dispenses", [])
        acc: Dict[str, float] = {}

        non_solvent_liquids = 0.0
        solvent_entries: List[Dict[str, Any]] = []

        for d in dispenses:
            group = str(d.get("group", "")).strip()
            name  = str(d.get("chemicalName", "")).strip()
            state = d.get("physicalstate")
            vol   = d.get("volume_uL")
            mass  = d.get("mass_mg")

            if group in _SOLVENT_GROUPS:
                solvent_entries.append(d)
                continue

            # liquids
            if vol is not None and (state in _LIQUID_STATES or state is None):
                try:
                    v = float(vol)
                    non_solvent_liquids += v
                    acc[f"{name} (uL)"] = acc.get(f"{name} (uL)", 0.0) + v
                except Exception:
                    pass

            # solids
            if mass is not None:
                try:
                    m = float(mass)
                    acc[f"{name} (mg)"] = acc.get(f"{name} (mg)", 0.0) + m
                except Exception:
                    pass

        # compute solvent remainder and split
        if V_target is not None:
            remainder = V_target - non_solvent_liquids
            if remainder < -1e-6:
                warns.append(
                    f"Liquids exceed target in Experiment {i} "
                    f"({non_solvent_liquids:.2f} > {V_target:.2f} µL); "
                    f"solvent set to 0."
                )
                remainder = 0.0

            if solvent_entries:
                # weights by 'ratio' (preferred) or 'equivalents'; else 1.0 each
                weights: List[float] = []
                for s in solvent_entries:
                    w = None
                    for key in ("ratio", "equivalents"):
                        if s.get(key) is not None:
                            try:
                                w = float(s.get(key))
                                break
                            except Exception:
                                pass
                    weights.append(float(w) if w is not None else 1.0)
                total_w = sum(weights) if sum(weights) > 0 else len(solvent_entries)
                for s, w in zip(solvent_entries, weights):
                    name = str(s.get("chemicalName", "")).strip()
                    share = (w / total_w) if total_w > 0 else (1.0 / len(solvent_entries))
                    v = remainder * share
                    acc[f"{name} (uL)"] = acc.get(f"{name} (uL)", 0.0) + v
            else:
                if remainder > 1e-6:
                    warns.append(
                        f"No solvent defined in Experiment {i}, but target "
                        f"implies {remainder:.2f} µL solvent."
                    )
        else:
            # No target: pass through any explicit solvent volumes (if given)
            for s in solvent_entries:
                name = str(s.get("chemicalName", "")).strip()
                vol  = s.get("volume_uL")
                if vol is not None:
                    try:
                        v = float(vol)
                        acc[f"{name} (uL)"] = acc.get(f"{name} (uL)", 0.0) + v
                    except Exception:
                        pass

        rows.append((f"Experiment {i}", acc))

    # dataframe
    all_cols = sorted(set().union(*[set(d.keys()) for _, d in rows])) if rows else []
    data = []
    idx = []
    for label, d in rows:
        idx.append(label)
        data.append([d.get(c, 0.0) for c in all_cols])
    df = pd.DataFrame(data, index=idx, columns=all_cols)
    info = {"V_target": V_target, "n_exp": len(order)}
    return df, info, warns

"""
-------- Actually filling excel workflow from synthesis --------
"""
_STOP_NODE_DEFAULT = r"(Orbital Shaker|Tumbler Stirrer|Magnetic Stirrer)"

def _find_header_row(ws) -> int:
    for r in range(1, 30):
        row_vals = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        if any(v.lower() == "wf node" for v in row_vals) and any(v.lower() == "label" for v in row_vals):
            return r
    return 8  # legacy fallback

def _find_stop_row(ws, header_row: int, stop_node_regex: str) -> int:
    pat = re.compile(stop_node_regex, re.IGNORECASE)
    cols = {str(ws.cell(row=header_row, column=c).value).strip().lower(): c
            for c in range(1, ws.max_column + 1)}
    wf_node_col = cols.get("wf node")
    if not wf_node_col:
        return 10**9
    r = header_row + 1
    while r <= ws.max_row:
        v = ws.cell(row=r, column=wf_node_col).value
        txt = (str(v).strip() if v is not None else "")
        if pat.search(txt):
            return r
        r += 1
    return 10**9

def fill_workflow_from_synthesis(
    synthesis_path: Path,
    workflow_path: Path,
    *,
    stop_node_regex: str = _STOP_NODE_DEFAULT,
    label_map_csv: Optional[Path] = None,
) -> List[str]:
    """
    Fills 'Experiment Definition' directly from synthesis.json.
    Returns a list of warnings (also printed to stderr).
    """
    synth = _load_synthesis(synthesis_path)
    per_exp_df, info, syn_warns = _per_experiment_from_synthesis(synth)

    wb = openpyxl.load_workbook(workflow_path)
    if "Experiment Definition" not in wb.sheetnames:
        raise RuntimeError("Sheet 'Experiment Definition' not found.")
    ws = wb["Experiment Definition"]

    header_row = _find_header_row(ws)
    stop_row   = _find_stop_row(ws, header_row, stop_node_regex)

    # columns
    cols = {str(ws.cell(row=header_row, column=c).value).strip(): c
            for c in range(1, ws.max_column + 1)}
    label_col = None
    for k, c in cols.items():
        if k.lower() == "label":
            label_col = c
            break
    if not label_col:
        raise RuntimeError("Column 'LABEL' not found in header.")

    # experiment columns
    exp_cols: Dict[str, int] = {}
    if not exp_cols:
        # (A) look a few rows ABOVE header for "Experiment " labels (common in ChemSpeed)
        for probe_row in range(1, header_row):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=probe_row, column=c).value
                if v and str(v).strip().lower().startswith("experiment "):
                    exp_cols[str(v).strip()] = c
        # (B) if still none: assume all columns to the RIGHT of "DEFAULT" are experiments
        if not exp_cols:
            cols_lc = {str(ws.cell(row=header_row, column=c).value).strip().lower(): c
                       for c in range(1, ws.max_column + 1)}
            default_col = cols_lc.get("default")
            if default_col:
                k = 1
                for c in range(default_col + 1, ws.max_column + 1):
                    exp_cols[f"Experiment {k}"] = c
                    k += 1
    if not exp_cols:
        print("⚠️  No experiment columns detected. Nothing to write.", file=sys.stderr)
    else:
        print(f"[info] Detected {len(exp_cols)} experiment column(s): "
              f"{', '.join(sorted(exp_cols.keys(), key=lambda x: int(x.split()[-1])))}")

    # normalize synthesis columns
    syn_cols_norm = {_norm_label(c): c for c in per_exp_df.columns}
    label_overrides = _load_label_map(label_map_csv)

    unknown_labels = set()
    r = header_row + 1
    while r < stop_row and r <= ws.max_row:
        raw_label = ws.cell(row=r, column=label_col).value
        if not raw_label or str(raw_label).strip() == "":
            r += 1
            continue

        wf_norm = _norm_label(str(raw_label))

        # override?
        syn_col = label_overrides.get(wf_norm)
        if syn_col is None:
            # exact normalized match first
            syn_col = syn_cols_norm.get(wf_norm)
        if syn_col is None:
            # startswith fallback
            matches = [orig for norm, orig in syn_cols_norm.items() if norm.startswith(wf_norm)]
            if len(matches) == 1:
                syn_col = matches[0]

        if syn_col is None:
            unknown_labels.add(str(raw_label))
            r += 1
            continue

        # write across experiments
        for exp_label, c in exp_cols.items():
            if exp_label not in per_exp_df.index:
                continue
            val = float(per_exp_df.loc[exp_label].get(syn_col, 0.0))
            ws.cell(row=r, column=c).value = val

        r += 1

    if unknown_labels:
        syn_warns.append("Unmatched workflow LABELs (no synthesis column found):")
        for lbl in sorted(unknown_labels):
            syn_warns.append(f"  - {lbl}")

    wb.save(workflow_path)
    print(f"[ok] Filled workflow from synthesis: {workflow_path}")
    for w in syn_warns:
        print(f"⚠️  {w}", file=sys.stderr)
    return syn_warns

# ==== VALIDATOR ====
def validate_synthesis_for_workflow(
    synthesis_path: Path,
    *,
    allow_tiny_neg: float = -1e-6
) -> List[str]:
    """
    Quick consistency checks for a synthesis that will feed the workflow:
      - Non-negative volumes/masses
      - Liquids do not exceed target well volume
      - Solvent present when remainder > 0 (or warn)
      - Plate size matches number of experiments
    Returns a list of warnings/errors (also printed).
    """
    synth = _load_synthesis(synthesis_path)
    meta = synth.get("meta", {})
    exps = synth.get("experiments", {})

    msgs: List[str] = []

    # plate size check
    plate_size_meta = None
    try:
        plate_size_meta = int(meta.get("layout", {}).get("plate_size") or 0)
    except Exception:
        pass
    n_exp = len(exps.keys())
    if plate_size_meta and plate_size_meta != n_exp:
        msgs.append(f"Plate size mismatch: meta.layout.plate_size={plate_size_meta} vs experiments={n_exp}")

    # target volume
    V_target = None
    lim = meta.get("limiting", {})
    if "well_volume_uL" in lim:
        try:
            V_target = float(lim["well_volume_uL"])
        except Exception:
            pass

    for eid, exp in exps.items():
        dispenses = exp.get("dispenses", [])
        non_solvent_liqs = 0.0
        solvents = 0
        for d in dispenses:
            name = d.get("chemicalName", "")
            vol  = d.get("volume_uL")
            mass = d.get("mass_mg")
            grp  = str(d.get("group", "")).strip()
            state = d.get("physicalstate")

            if vol is not None:
                try:
                    v = float(vol)
                    if v < allow_tiny_neg:
                        msgs.append(f"Negative volume in exp {eid} for {name}: {v} µL")
                    if grp not in _SOLVENT_GROUPS and (state in _LIQUID_STATES or state is None):
                        non_solvent_liqs += v
                except Exception:
                    msgs.append(f"Non-numeric volume in exp {eid} for {name}: {vol}")

            if mass is not None:
                try:
                    m = float(mass)
                    if m < allow_tiny_neg:
                        msgs.append(f"Negative mass in exp {eid} for {name}: {m} mg")
                except Exception:
                    msgs.append(f"Non-numeric mass in exp {eid} for {name}: {mass}")

            if grp in _SOLVENT_GROUPS:
                solvents += 1

        if V_target is not None:
            remainder = V_target - non_solvent_liqs
            if remainder < -1e-6:
                msgs.append(
                    f"Liquids exceed target in exp {eid}: {non_solvent_liqs:.2f} > {V_target:.2f} µL"
                )
            if remainder > 1e-6 and solvents == 0:
                msgs.append(
                    f"Missing solvent in exp {eid}: remainder {remainder:.2f} µL (no solvent/colvents group)"
                )

    if msgs:
        for m in msgs:
            print(f"⚠️  {m}", file=sys.stderr)
    else:
        print("[ok] synthesis passed basic validation.", file=sys.stderr)
    return msgs

"""
-------------- New Visuals --------------
"""
def _plate_dims_from_meta(meta: dict) -> tuple[int, int]:
    """Infer (rows, cols) from meta.layout if present; else from plate_size."""
    lay = meta.get("layout", {}) or {}
    rows = int(lay.get("rows") or 0)
    cols = int(lay.get("cols") or 0)
    if rows and cols:
        return rows, cols
    # fallback from plate_size
    ps = int(lay.get("plate_size") or meta.get("limiting", {}).get("plate_size") or 0)
    # simple heuristics
    if ps in (96,):  return 8, 12
    if ps in (48,):  return 6, 8
    if ps in (24,):  return 4, 6
    if ps in (18,):  return 3, 6
    if ps in (12,):  return 3, 4
    if ps in (6,):   return 2, 3
    # last resort: square-ish
    side = int(math.ceil(math.sqrt(ps or 0)))
    return side, side


def _experiment_grid_from_synthesis(synth: dict) -> tuple[list[list[str | None]], list[str]]:
    """
    Returns:
      grid  -> 2D list [rows][cols] with 'Experiment N' labels or None
      order -> list of experiment IDs as strings in the same N order
    """
    meta = synth.get("meta", {})
    exps = synth.get("experiments", {})
    # order: honor meta.experiment_to_well if available
    exp_order = [str(x["experiment"]) for x in (meta.get("experiment_to_well") or [])]
    if not exp_order:
        exp_order = sorted(exps.keys(), key=lambda z: int(str(z)))

    rows, cols = _plate_dims_from_meta(meta)
    grid = [[None for _ in range(cols)] for _ in range(rows)]

    # place by well if given, else left-to-right fill
    well_map = {str(x["experiment"]): str(x.get("well") or "") for x in (meta.get("experiment_to_well") or [])}

    def _rc_from_well(w: str) -> tuple[int, int] | None:
        # Converts e.g. 'A1'.. to 0-based (row, col). Return None if cannot parse.
        if not w or not w[0].isalpha():
            return None
        r = ord(w[0].upper()) - ord('A')
        try:
            c = int(w[1:]) - 1
        except Exception:
            return None
        return (r, c) if 0 <= r < rows and 0 <= c < cols else None

    # fill by mapping if present
    used = set()
    n = 0
    for eid in exp_order:
        n += 1
        label = f"Experiment {n}"
        rc = _rc_from_well(well_map.get(eid, ""))
        if rc:
            r, c = rc
            grid[r][c] = label
            used.add(eid)

    # place remaining (if any) left-to-right
    for eid in exp_order:
        if eid in used:
            continue
        # find first None slot
        placed = False
        for r in range(rows):
            for c in range(cols):
                if grid[r][c] is None:
                    n = len([cell for row in grid for cell in row if cell is not None]) + 1
                    grid[r][c] = f"Experiment {n}"
                    placed = True
                    break
            if placed: break
    return grid, [f"Experiment {i+1}" for i in range(len(exp_order))]

def _chemical_states_from_synthesis(experiments: dict) -> dict[str, str]:
    """
    Build a mapping chemicalName -> physicalstate (lowercased),
    taken from the synthesis.json (top level, reference or reference.extras).
    """
    states: dict[str, str] = {}
    for e in experiments.values():
        for d in e.get("dispenses", []):
            name = str(d.get("chemicalName", "")).strip()
            if not name or name in states:
                continue

            state = (
                d.get("physicalstate")
                or d.get("reference", {}).get("physicalstate")
                or d.get("reference", {}).get("extras", {}).get("physicalstate")
            )
            if isinstance(state, str) and state.strip():
                states[name] = state.strip().lower()
    return states

def visualize_plate_from_synthesis(
    synthesis_path: Path,
    out_png: str,
    *,
    title: str = "Experiment Map",
) -> None:
    """Simple per-well map with experiment numbers drawn on each well."""
    synth = _load_synthesis(Path(synthesis_path))
    grid, _ = _experiment_grid_from_synthesis(synth)

    rows, cols = len(grid), len(grid[0]) if grid else (0, 0)
    fig, ax = plt.subplots(figsize=(cols * 0.5 + 1.5, rows * 0.5 + 1.5), dpi=200)
    ax.imshow([[0 if cell is None else 1 for cell in row] for row in grid])
    for i in range(rows):
        for j in range(cols):
            label = grid[i][j]
            if label:
                token = label.split()[-1] if " " in label else str(label)
                ax.text(j, i, token, ha='center', va='center', color='black', fontsize=8)
    ax.set_title(title)
    ax.set_xticks(range(cols)); ax.set_yticks(range(rows))
    # nicer axis labels A.. and 1..N
    ax.set_yticklabels([chr(ord('A') + r) for r in range(rows)])
    ax.set_xticklabels([str(c+1) for c in range(cols)])
    ax.set_xlabel("Columns"); ax.set_ylabel("Rows")
    ax.grid(False)
    plt.tight_layout()
    fig.savefig(out_png, bbox_inches="tight")
    plt.close(fig)


def visualize_heatmap_for_chemicals(
    synthesis_path: Path,
    out_dir: str,
    chemicals: Iterable[str] | None = None,
    use_seaborn: bool = True,
) -> list[str]:
    """
    For each chosen chemical (or for all chemicals if None),
    draw a plate heatmap (uL or mg) using the synthesis.json content.
    Returns list of saved PNG paths.
    """
    synth = _load_synthesis(Path(synthesis_path))
    grid, exp_labels = _experiment_grid_from_synthesis(synth)
    meta = synth.get("meta", {})
    exps = synth.get("experiments", {})

    rows, cols = len(grid), len(grid[0]) if grid else (0, 0)

    # Map exp_label -> (row,col)
    rc_by_exp = {}
    for i in range(rows):
        for j in range(cols):
            if grid[i][j]:
                rc_by_exp[grid[i][j]] = (i, j)

    # Collect all chemicalNames present
    all_names: set[str] = set()
    for e in exps.values():
        for d in e.get("dispenses", []):
            nm = str(d.get("chemicalName", "")).strip()
            if nm:
                all_names.add(nm)

    chem_states = _chemical_states_from_synthesis(exps)
    picks = list(chemicals) if chemicals else sorted(all_names)

    saved: list[str] = []
    for name in picks:
        state = chem_states.get(name, "").lower()

        # You can tweak these sets if you introduce more labels
        is_solid = state in {"solid", "powder", "solid (powder)"}
        is_liquid = state in {"liquid", "solution"}

        if is_solid:
            field_key = "mass_mg"
            unit = "mg"
            cbar_label = "Mass (mg)"
        else:
            # default to liquid behaviour if unknown
            field_key = "volume_uL"
            unit = "uL"
            cbar_label = "Volume (uL)"
            if not is_liquid and state:
                # Unknown state string → warn but still plot volume
                print(
                    f"⚠️  Unknown physicalstate '{state}' for '{name}', "
                    f"defaulting to volume_uL.",
                    file=sys.stderr,
                )
        Z = [[0.0 for _ in range(cols)] for _ in range(rows)]
        for idx, exp_id in enumerate(sorted(exps.keys(), key=lambda z: int(str(z))), start=1):
            label = f"Experiment {idx}"
            cell = rc_by_exp.get(label)
            if not cell:
                continue
            i, j = cell
            # sum volume or mass for this chemical in this experiment
            total = 0.0
            for d in exps[str(exp_id)].get("dispenses", []):
                if str(d.get("chemicalName", "")).strip() != name:
                    continue
                val = d.get(field_key)
                if val is not None:
                    try:
                        total += float(val)
                    except Exception:
                        pass
            Z[i][j] = total

        if _HAS_SNS and use_seaborn:
            # seaborn path
            fig, ax = plt.subplots(figsize=(cols * 0.5 + 1.5, rows * 0.5 + 1.5), dpi=200)
            # We want (rows x cols) matrix; seaborn expects a 2D array
            sns.heatmap(
                np.array(Z),
                ax=ax,
                annot=False,
                cbar=True,
                square=True,
                linewidths=0.3,
                linecolor="white",
            )
            cbar = ax.collections[0].colorbar
            cbar.set_label(f"{'Amount' if 'solvent' not in locals() else 'Volume'} ({unit})", rotation=270, labelpad=15)

            ax.set_title(f"{name} — {unit}")
            ax.set_xticks(np.arange(cols) + 0.5)
            ax.set_yticks(np.arange(rows) + 0.5)
            ax.set_xticklabels([str(c + 1) for c in range(cols)], rotation=0)
            ax.set_yticklabels([chr(ord('A') + r) for r in range(rows)], rotation=0)
        else:
            # original matplotlib path
            fig, ax = plt.subplots(figsize=(cols * 0.5 + 1.5, rows * 0.5 + 1.5), dpi=200)
            im = ax.imshow(Z, cmap="viridis")  # or 'plasma', 'coolwarm', etc.
            cbar = fig.colorbar(im, ax=ax, shrink=0.8, pad=0.02)
            cbar.set_label(f"{'Amount' if 'solvent' not in locals() else 'Volume'} ({unit})", rotation=270, labelpad=15)
            ax.set_title(f"{name} — {unit}" if 'name' in locals() else f"Solvent remainder — {unit}")
            ax.set_xticks(range(cols))
            ax.set_yticks(range(rows))
            ax.set_yticklabels([chr(ord('A') + r) for r in range(rows)])
            ax.set_xticklabels([str(c + 1) for c in range(cols)])
            ax.grid(False)


        plt.tight_layout()
        out_png = str(Path(out_dir) / f"plate_heatmap_{name.replace(' ','_')}_{unit}.png")
        fig.savefig(out_png, bbox_inches="tight")
        plt.close(fig)
        saved.append(out_png)

    return saved

def visualize_heatmap_solvent_remainder(
    synthesis_path: Path,
    out_dir: str,
    use_seaborn: bool = False,
) -> str:
    """
    Heatmap of remaining solvent per well:
    remainder_uL = V_target - sum(non-solvent liquid volumes)
    (clamped at 0 if negative).
    """
    synth = _load_synthesis(Path(synthesis_path))
    grid, exp_labels = _experiment_grid_from_synthesis(synth)
    meta = synth.get("meta", {})
    exps = synth.get("experiments", {})

    rows, cols = len(grid), len(grid[0]) if grid else (0, 0)
    # map 'Experiment N' -> (row,col)
    rc_by_exp = {}
    for i in range(rows):
        for j in range(cols):
            if grid[i][j]:
                rc_by_exp[grid[i][j]] = (i, j)

    # target volume
    V_target = None
    lim = meta.get("limiting", {})
    if "well_volume_uL" in lim:
        try:
            V_target = float(lim["well_volume_uL"])
        except Exception:
            V_target = None

    Z = [[0.0 for _ in range(cols)] for _ in range(rows)]
    if V_target is not None:
        # iterate experiments in numeric order to match Experiment 1..N labels
        for idx, exp_id in enumerate(sorted(exps.keys(), key=lambda z: int(str(z))), start=1):
            label = f"Experiment {idx}"
            cell = rc_by_exp.get(label)
            if not cell:
                continue
            i, j = cell
            # sum non-solvent liquids
            non_solvent_liqs = 0.0
            for d in exps[str(exp_id)].get("dispenses", []):
                grp = str(d.get("group", "")).strip().lower()
                vol = d.get("volume_uL")
                state = d.get("physicalstate")
                if grp in _SOLVENT_GROUPS:
                    continue
                if vol is not None and (state in _LIQUID_STATES or state is None):
                    try:
                        non_solvent_liqs += float(vol)
                    except Exception:
                        pass
            rem = V_target - non_solvent_liqs
            if rem < 0:
                rem = 0.0
            Z[i][j] = rem

    # plot
    out_png = str(Path(out_dir) / "plate_heatmap_solvent_remainder_uL.png")
    if use_seaborn and _HAS_SNS:
        import numpy as np
        fig, ax = plt.subplots(figsize=(cols * 0.5 + 1.5, rows * 0.5 + 1.5), dpi=200)
        sns.heatmap(
            np.array(Z),
            ax=ax,
            annot=False,
            cbar=True,
            square=True,
            linewidths=0.3,
            linecolor="white",
        )
        ax.set_title("Solvent remainder — uL")
        cbar = ax.collections[0].colorbar
        cbar.set_label("Remaining solvent volume (uL)", rotation=270, labelpad=15)
        ax.set_xticks(np.arange(cols) + 0.5)
        ax.set_yticks(np.arange(rows) + 0.5)
        ax.set_xticklabels([str(c+1) for c in range(cols)], rotation=0)
        ax.set_yticklabels([chr(ord('A') + r) for r in range(rows)], rotation=0)
        plt.tight_layout()
        fig.savefig(out_png, bbox_inches="tight")
        plt.close(fig)
    else:
        fig, ax = plt.subplots(figsize=(cols * 0.5 + 1.5, rows * 0.5 + 1.5), dpi=200)
        im = ax.imshow(Z, cmap="viridis")  # or 'plasma', 'coolwarm', etc.
        cbar = fig.colorbar(im, ax=ax, shrink=0.8, pad=0.02)
        unit = "uL"
        cbar.set_label(f"{'Amount' if 'solvent' not in locals() else 'Volume'} ({unit})", rotation=270, labelpad=15)
        ax.set_title(f"Solvent remainder — {unit}")
        ax.set_xticks(range(cols))
        ax.set_yticks(range(rows))
        ax.set_yticklabels([chr(ord('A') + r) for r in range(rows)])
        ax.set_xticklabels([str(c + 1) for c in range(cols)])
        ax.grid(False)
        plt.tight_layout()
        fig.savefig(out_png, bbox_inches="tight")
        plt.close(fig)

    return out_png



"""
-------------- Legacy code passing via calculator excel --------------
"""

def parse_calculator_excel(path: str) -> Tuple[List[str], List[str], List[str]]:
    """Return lists of solid reagents, liquid reagents and solvents from
    an HTE calculator output file."""
    xl = pd.ExcelFile(path)
    df = pd.read_excel(xl, "per_well", header=None)
    headers = [h for h in df.iloc[0, 1:] if isinstance(h, str)]
    unique = list(dict.fromkeys(headers))

    solids: List[str] = []
    liquids: List[str] = []
    solvents: List[str] = []

    for name in unique:
        low = name.lower()
        if "as stock solution" in low:
            # ignore amounts needed to prepare stock solutions
            continue
        if "(mg)" in low:
            solids.append(name)
            continue
        if " in " in name:
            liquids.append(name)
            continue
        if "(ul)" in low or "microliter" in low:
            if name.count(" ") == 1:
                solvents.append(name)
            else:
                liquids.append(name)
    return solids, liquids, solvents


def workflow_dispense_counts(path: str) -> Tuple[int, int, List[str]]:
    """Return number of solid and liquid dispense steps before the first
    orbital shaker as well as the list of workflow steps."""
    df, _ = lp.read_experiment_definition(path)
    steps = lp.extract_workflow_steps(df)

    orbital_idx = df.index[df["WF NODE"].str.contains("Orbital Shaker", na=False)]
    first_orbital = orbital_idx.min() if not orbital_idx.empty else len(df)

    pre = df.loc[: first_orbital - 1]
    solids = pre[(pre["TYPE"] == "Product") & pre["UNIT"].str.contains("gram|milligram", case=False, na=False)]
    liquids = pre[(pre["TYPE"] == "Product") & pre["UNIT"].str.contains("microliter", case=False, na=False)]

    return len(solids), len(liquids), steps


def per_well_to_experiments(
    per_well: pd.DataFrame, plate_layout: pd.DataFrame
) -> pd.DataFrame:
    """Return per-experiment reagent amounts from calculator output."""
    exp_index = [label for label in plate_layout.stack()]
    result = pd.DataFrame(index=exp_index)
    for reagent in per_well.columns.levels[0]:
        sub = per_well[reagent]
        values: Dict[str, float] = {}
        for row in sub.index:
            for col in sub.columns:
                exp = plate_layout.loc[row, col]
                values[exp] = float(sub.loc[row, col])
        result[reagent] = pd.Series(values)
    # ensure experiments are sorted numerically
    result.index = pd.CategoricalIndex(
        result.index, ordered=True, categories=sorted(result.index, key=lambda x: int(x.split()[1]))
    )
    result.sort_index(inplace=True)
    return result


def fill_workflow(calculator: str, workflow: str) -> None:
    """Fill an empty workflow file with amounts from the calculator export.

    The workflow file is modified in place."""
    wf_df, mapping = lp.read_experiment_definition(workflow)
    num_exp = len(mapping)

    per = pd.read_excel(calculator, "per_well", header=[0, 1], index_col=0)
    plate_layout, _ = lp.map_experiments_to_wells(num_exp)
    per_exp = per_well_to_experiments(per, plate_layout)

    wb = openpyxl.load_workbook(workflow)
    ws = wb["Experiment Definition"]

    rev_map = {v: k + 1 for k, v in mapping.items()}

    orbital_idx = wf_df.index[wf_df["WF NODE"].str.contains("Orbital Shaker", na=False)]
    first_orbital = orbital_idx.min() if not orbital_idx.empty else len(wf_df)

    header_offset = 8  # number of header rows before pandas data

    for idx, row in wf_df.loc[: first_orbital - 1].iterrows():
        label = row.get("LABEL")
        if not isinstance(label, str):
            continue

        calc_name = None
        if label in per_exp.columns:
            calc_name = label
        else:
            for col in per_exp.columns:
                if isinstance(col, str) and col.startswith(label):
                    calc_name = col
                    break

        if not calc_name:
            continue

        excel_row = idx + header_offset + 1
        for exp_label, col_idx in rev_map.items():
            val = per_exp.at[exp_label, calc_name]
            ws.cell(row=excel_row, column=col_idx, value=val)

    wb.save(workflow)


"""
-------------- End of legacy code --------------
"""


def main() -> None:
    parser = argparse.ArgumentParser(description="Check workflow against calculator output")
    parser.add_argument("calculator", help="Excel file from hte_calculator")
    parser.add_argument("workflow", help="Workflow Excel file (empty)")
    parser.add_argument("--visualize", action="store_true", help="Run layout_parser to generate images")
    parser.add_argument(
        "--fill",
        action="store_true",
        help="Write dispense amounts directly into the workflow file",
    )
    parser.add_argument("--fill-from-synthesis", action="store_true",
                        help="Fill workflow directly from a synthesis.json.")
    parser.add_argument("--synthesis-file", help="synthesis.json (required for --fill-from-synthesis).")
    parser.add_argument("--stop-node", default=_STOP_NODE_DEFAULT,
                        help="Regex for first node to stop at (default matches Shaker/Stirrer).")
    parser.add_argument("--label-map-csv", help="Optional CSV remapping 'workflow_label,synthesis_name'.")
    parser.add_argument("--validate-synthesis", action="store_true", help="Run validator on the synthesis and exit.")

    # For visuals
    parser.add_argument("--viz-plate", action="store_true",
                        help="Save a plate map PNG from synthesis.json.")
    parser.add_argument("--viz-heatmaps", action="store_true",
                        help="Save per-chemical heatmaps from synthesis.json.")
    parser.add_argument("--viz-chemicals", nargs="*", default=None,
                        help="Subset of chemical names to plot (default: all).")
    parser.add_argument("--viz-seaborn", action="store_true", default = True,
                        help="Use seaborn styling for heatmaps (if available).") # needs change that False is treated correctly
    parser.add_argument("--viz-solvent-remainder", action="store_true",
                        help="Save a solvent remainder heatmap (target − other liquids, uL).")

    # Data and output directories

    parser.add_argument("--data-dir",
                        default=str(DATA_DIR),
                        help="Directory with data files for layout_parser")
    parser.add_argument("--out-dir",
                        default=str(OUT_DIR),
                        help="Directory for output files from layout_parser")
    args = parser.parse_args()

    data_dir = Path(args.data_dir).resolve()
    out_dir = Path(args.out_dir).resolve()

    calculator_excel_path = out_dir / args.calculator
    workflow_excel_path = data_dir / args.workflow
    synthesis_file_path = Path(out_dir / args.synthesis_file) if args.synthesis_file else None

    if args.validate_synthesis:
        if not args.synthesis_file:
            raise SystemExit("--validate-synthesis needs --synthesis-file")
        validate_synthesis_for_workflow(synthesis_file_path)
        sys.exit(0)

    if args.fill_from_synthesis:
        if not args.synthesis_file or not args.workflow:
            raise SystemExit("--fill-from-synthesis needs --synthesis-file and --workflow")
        fill_workflow_from_synthesis(
            synthesis_file_path,
            workflow_excel_path,
            stop_node_regex=args.stop_node,
            label_map_csv=Path(args.label_map_csv) if args.label_map_csv else None,
        )
        # Save visuals if requested
        if args.viz_plate or args.viz_heatmaps or args.viz_solvent_remainder:
            vis_dir = out_dir / (Path(synthesis_file_path).stem + "_visuals")
            vis_dir.mkdir(parents=True, exist_ok=True)

            if args.viz_plate:
                png = str(vis_dir/Path(workflow_excel_path).stem) + "_plate_map.png"
                visualize_plate_from_synthesis(synthesis_file_path, png, title="Experiment Map")
                print(f"[ok] saved plate map → {png}")

            if args.viz_heatmaps:
                saved = visualize_heatmap_for_chemicals(synthesis_file_path, str(vis_dir), chemicals=args.viz_chemicals,
                                                        use_seaborn=args.viz_seaborn)
                if saved:
                    print("[ok] saved heatmaps:")
                    for p in saved:
                        print("   ", p)

            if args.viz_solvent_remainder:
                png = visualize_heatmap_solvent_remainder(
                    synthesis_file_path,
                    str(vis_dir),
                    use_seaborn=args.viz_seaborn,
                )
                print(f"[ok] saved solvent remainder heatmap → {png}")
        sys.exit(0)

    calc_solids, calc_liquids, calc_solvents = parse_calculator_excel(calculator_excel_path)
    wf_solids, wf_liquids, steps = workflow_dispense_counts(workflow_excel_path)

    print(f"Calculator reagents: {len(calc_solids)} solids, {len(calc_liquids)} liquids, {len(calc_solvents)} solvents")
    print(f"Workflow before first orbital shaker has {wf_solids} solid dispenses and {wf_liquids} liquid dispenses")
    if len(calc_solids) != wf_solids:
        print("Warning: mismatch in number of solid dispenses")
    if len(calc_liquids) + len(calc_solvents) != wf_liquids:
        print("Warning: mismatch in number of liquid/solvent dispenses")

    if args.fill:
        fill_workflow(calculator_excel_path, workflow_excel_path)
        print(f"Workflow {args.workflow} updated")

    if args.visualize:
        # run layout_parser with the workflow file only so its own
        # argument parser does not see the workflow checker's options

        original = sys.argv
        sys.argv = [
                    "lp",
                    "--data-dir", str(data_dir),
                    "--out-dir", str(out_dir),
                    args.workflow]
        try:
            lp.main()
        finally:
            sys.argv = original


if __name__ == "__main__":
    main()
